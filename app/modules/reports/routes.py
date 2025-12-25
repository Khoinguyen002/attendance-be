from io import BytesIO
from bson import ObjectId
from flask import Blueprint, request, send_file

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.common.decorators import role_required
from app.extensions.mongo import mongo

report_bp = Blueprint("reports", __name__)


@report_bp.get("/daily")
@role_required("manager")
def report_daily():
    date = request.args.get("date")
    if not date:
        return {"message": "date is required (YYYY-MM-DD)"}, 400

    records = list(
        mongo.db.attendance_daily.find(
            {"date": date},
            {"_id": 0}
        )
    )

    for r in records:
        r["employee_id"] = str(r["employee_id"])

    return {
        "date": date,
        "total": len(records),
        "records": records
    }, 200


@report_bp.get("/monthly")
@role_required("manager")
def report_monthly():
    month = request.args.get("month")
    if not month:
        return {"message": "month is required (YYYY-MM)"}, 400

    records = list(
        mongo.db.attendance_daily.find(
            {"date": {"$regex": f"^{month}"}},
            {"_id": 0}
        )
    )

    for r in records:
        r["employee_id"] = str(r["employee_id"])

    return {
        "month": month,
        "total": len(records),
        "records": records
    }, 200


@report_bp.get("/employee/<employee_id>/monthly")
@role_required("manager")
def report_employee_monthly(employee_id):
    month = request.args.get("month")
    if not month:
        return {"message": "month is required (YYYY-MM)"}, 400

    records = list(
        mongo.db.attendance_daily.find(
            {
                "employee_id": ObjectId(employee_id),
                "date": {"$regex": f"^{month}"}
            },
            {"_id": 0}
        )
    )

    for r in records:
        r["employee_id"] = str(r["employee_id"])

    return {
        "employee_id": employee_id,
        "month": month,
        "total": len(records),
        "records": records
    }, 200

@report_bp.get("/export/monthly")
@role_required("manager")
def export_monthly():
    month = request.args.get("month")
    if not month:
        return {"message": "month is required (YYYY-MM)"}, 400

    records = list(
        mongo.db.attendance_daily.find(
            {"date": {"$regex": f"^{month}"}},
            {"_id": 0}
        )
    )

    if not records:
        return {"message": "No data"}, 404

    # ==== LOAD EMPLOYEE MAP ====
    employee_ids = list({r["employee_id"] for r in records})
    employees = mongo.db.employees.find(
        {"_id": {"$in": employee_ids}},
        {"name": 1, "email": 1}
    )
    emp_map = {
        e["_id"]: {
            "name": e.get("name"),
            "email": e.get("email")
        }
        for e in employees
    }

    # ==== CREATE EXCEL ====
    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance {month}"

    headers = [
        "Date",
        "Employee ID",
        "Employee Name",
        "Email",
        "Status",
        "Late (min)",
        "Early (min)",
        "Worked (min)"
    ]
    ws.append(headers)

    for r in records:
        emp = emp_map.get(r["employee_id"], {})
        ws.append([
            r.get("date"),
            str(r.get("employee_id")),
            emp.get("name"),
            emp.get("email"),
            r.get("status"),
            r.get("late_minutes", 0),
            r.get("early_minutes", 0),
            r.get("worked_minutes", 0),
        ])

    # Auto width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # ==== SAVE TO MEMORY ====
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"attendance_{month}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )